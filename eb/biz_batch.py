# coding: UTF-8
"""
Created on 2016/06/03

@author: Yang Wanjun
"""
import datetime
import urllib2
import json
import re
import os
import traceback
import openpyxl as px

from . import biz, biz_config
from utils import constants, common, file_gen
from eb import models
from eboa import models as eboa_models

from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE
from django.contrib.contenttypes.models import ContentType
from django.utils.text import get_text_list
from django.utils.translation import ugettext as _


def sync_members(batch):
    company = biz.get_company()
    logger = batch.get_logger()
    url = biz_config.get_config(constants.CONFIG_SERVICE_MEMBERS)
    response = urllib2.urlopen(url)
    html = response.read()
    dict_data = json.loads(html.replace("\r", "").replace("\n", ""))
    count = 0
    user = batch.get_log_entry_user()
    if 'employeeList' in dict_data:
        for data in dict_data.get("employeeList"):
            # ＩＤ主キー
            api_id = data.get("id", None)
            # 名前
            name = data.get("name", None)
            # カタカナ
            name_jp = data.get("kana", None)
            first_name_ja = last_name_ja = None
            if name_jp:
                lst = common.get_first_last_ja_name(name_jp)
                if len(lst) == 2 and lst[0]:
                    first_name_ja = common.get_first_last_ja_name(name_jp)[0]
                    last_name_ja = common.get_first_last_ja_name(name_jp)[1]
                elif len(lst) == 1:
                    first_name_ja = common.get_first_last_ja_name(name_jp)[0]
            # 誕生日
            birthday = data.get("birthDate", None)
            if birthday:
                try:
                    birthday = common.parse_date_from_string(birthday)
                except Exception as ex:
                    logger.warning(u"%sの生年月日(%s)が取得できません。%s" % (name, birthday, unicode(ex)))
            else:
                birthday = None
            # 郵便番号
            postcode = data.get("postcode", None)
            if postcode and postcode.strip():
                postcode = postcode.strip().replace("/", "").replace("-", "").replace(u"‐", "").strip()
                if not re.match(r"[0-9]{7}", postcode):
                    postcode = None
                    logger.warning(u"%sの郵便番号(%s)が取得できません。" % (name, postcode))
            # 住所
            address = data.get("address", None)
            # 個人メールアドレス
            private_mail = data.get("mailAddress", None)
            # 会社メールアドレス
            eb_mail = data.get("ebMailAddress", None)
            # 携帯番号
            phone = data.get("phone", None)
            if phone:
                phone = phone.replace("-", "")
                if not re.match(r"[0-9]+", phone):
                    phone = None
                    logger.warning(u"%sの携帯番号(%s)が取得できません。" % (name, phone))
            # 性別
            sex = data.get("sex", None)
            # 最寄り駅
            station = data.get("station", None)
            # 入社年月日
            join_date = data.get("joinDate", None)
            if join_date:
                join_date = common.parse_date_from_string(join_date)
            # 給料王ＩＤ
            salary_id = data.get("salaryId", None)
            if salary_id and re.match(r"[0-9]+", salary_id):
                salary_id = "%06d" % int(salary_id)
            else:
                salary_id = api_id
            # 備考
            introduction = data.get("introduction", None)
            # 部署
            department_id = data.get('departmentId', None)
            department_name = data.get("department", None)
            org_id = constants.DICT_ORG_MAPPING.get(department_id, None)
            if api_id:
                # 契約情報取得する。
                # concat = get_latest_concat(api_id)
                # コスト
                # cost = get_cost(api_id)
                # if cost == 0:
                #     continue
                # if re.match(r"[0-9]+", str(cost)):
                #     cost = int(cost)
                # 社員区分
                # member_type = concat.get('EMPLOYER_TYPE', 0)
                # if member_type == u"正社員":
                #     member_type = 1
                # elif member_type == u"契約社員":
                #     member_type = 2
                # elif member_type == u"個人事業主":
                #     member_type = 3
                # else:
                #     logger.warning(u"%sの契約形態（%s）が識別できません。" % (name, member_type))
                #     member_type = 0
                try:
                    models.Member.objects.get(id_from_api=api_id)
                except ObjectDoesNotExist:
                    member = models.Member(employee_id=salary_id, id_from_api=api_id)
                    member.first_name = common.get_first_last_name(name)[0]
                    member.last_name = common.get_first_last_name(name)[1]
                    member.first_name_ja = first_name_ja
                    member.last_name_ja = last_name_ja
                    member.birthday = birthday
                    member.post_code = postcode
                    member.address1 = address
                    member.private_email = private_mail
                    member.email = eb_mail
                    member.phone = phone
                    member.sex = "2" if sex == "0" else "1"
                    member.nearest_station = station if station and len(station) <= 15 else None
                    member.join_date = join_date
                    member.comment = introduction
                    member.company = company
                    if department_id not in ('4', '5', '6', '7', '13', '18', '19', '20', '21', '22') and org_id:
                        try:
                            section = models.Section.objects.get(pk=org_id)
                            member.section = section
                        except ObjectDoesNotExist:
                            pass
                    member.save()
                    # ログ出力
                    change_message = _('Added.')
                    prefix = u"【%s】" % batch.title
                    LogEntry.objects.log_action(user_id=user.pk,
                                                content_type_id=ContentType.objects.get_for_model(member).pk,
                                                object_id=member.pk,
                                                object_repr=unicode(member),
                                                action_flag=ADDITION,
                                                change_message=(prefix + change_message) or _('No fields changed.'))

                    args = (member.employee_id, member.__unicode__(), u"追加完了。")
                    msg = u"社員コード: %s, name: %s, %s" % args
                    logger.info(msg)
                    count += 1
                except MultipleObjectsReturned:
                    # 一意制約なので、存在しないケース
                    pass
                except Exception as ex:
                    args = (salary_id, name, unicode(ex))
                    msg = u"社員コード: %s, name: %s, 予期しないエラー: %s" % args
                    logger.error(msg)
                    logger.error(traceback.format_exc())

    return count


def sync_members_for_change(batch):
    """EBOAのＤＢから社員の情報を更新する。

    :param batch:
    :return:
    """
    members = models.Member.objects.public_filter(eboa_user_id__isnull=False)
    logger = batch.get_logger()
    user = batch.get_log_entry_user()
    for member in members:
        try:
            oa_member = eboa_models.EbEmployee.objects.get(user__userid=member.eboa_user_id)
            changed_list = []
            zip_code = oa_member.zipcode.replace('-', '') if oa_member.zipcode else ""
            # 郵便番号
            if re.match(r'^[0-9]{7}$', zip_code) and member.post_code != zip_code:
                common.get_object_changed_message(member, 'post_code', zip_code, changed_list)
                member.post_code = zip_code
            # 住所
            if isinstance(oa_member.address, unicode):
                address = oa_member.address if oa_member.address else ''
            else:
                address = unicode(oa_member.address.decode('utf8')) if oa_member.address else ''
            old_address = member.address1 if member.address1 else ''
            old_address += member.address2 if member.address2 else ''
            if address and old_address != address:
                common.get_object_changed_message(member, 'address1', address, changed_list)
                common.get_object_changed_message(member, 'address2', "", changed_list)
                member.address1 = address
                member.address2 = ""
            # 電話番号
            private_tel_number = oa_member.private_tel_number.replace("-", "") if oa_member.private_tel_number else ''
            if re.match(r'^[0-9]{11}$', private_tel_number) and member.phone != private_tel_number:
                common.get_object_changed_message(member, 'phone', private_tel_number, changed_list)
                member.phone = private_tel_number
            # 会社メールアドレス
            if oa_member.business_mail_addr and oa_member.business_mail_addr.endswith("@e-business.co.jp") \
                    and member.email != oa_member.business_mail_addr:
                common.get_object_changed_message(member, 'email', oa_member.business_mail_addr, changed_list)
                member.email = oa_member.business_mail_addr
            if changed_list:
                member.save()
            if changed_list and user:
                change_message = _('Changed %s.') % get_text_list(changed_list, _('and')) if changed_list else ''
                prefix = u"【%s】" % batch.title
                LogEntry.objects.log_action(user_id=user.pk,
                                            content_type_id=ContentType.objects.get_for_model(member).pk,
                                            object_id=member.pk,
                                            object_repr=unicode(member),
                                            action_flag=CHANGE,
                                            change_message=(prefix + change_message) or _('No fields changed.'))
                args = (member.eboa_user_id, member.__unicode__(), u"情報が変更されました。")
                msg = u"eboa_user_id: %s, name: %s, %s" % args
                logger.info(msg)
        except ObjectDoesNotExist:
            args = (member.eboa_user_id, member.__unicode__(), u"ＥＢＯＡのＤＢに該当するデータがありません。")
            msg = u"eboa_user_id: %s, name: %s, %s" % args
            logger.warning(msg)
        except MultipleObjectsReturned:
            args = (member.eboa_user_id, member.__unicode__(), u"ＥＢＯＡのＤＢに該当するデータ複数存在している。")
            msg = u"eboa_user_id: %s, name: %s, %s" % args
            logger.warning(msg)
        except Exception as ex:
            args = (member.eboa_user_id, member.__unicode__(), unicode(ex))
            msg = u"eboa_user_id: %s, name: %s, 予期しないエラー: %s" % args
            logger.error(msg)
            logger.error(traceback.format_exc())
    # EBOAのユーザーＩＤが設定されてないメンバーを出力する。
    no_user_members = models.Member.objects.public_filter(eboa_user_id__isnull=True).exclude(member_type=4)
    name_list = []
    for member in no_user_members:
        name_list.append(u"%s(%s)" % (unicode(member), member.employee_id))
    if name_list:
        logger.warning(u"%sのEBOAユーザーＩＤが設定されていません。" % (', '.join(name_list)))


def sync_contracts(batch):
    members = models.Member.objects.public_filter(eboa_user_id__isnull=False)
    logger = batch.get_logger()
    user = batch.get_log_entry_user()
    # changed_list = []
    # common.get_object_changed_message(member, 'cost', cost, changed_list)
    # # common.get_object_changed_message(member, 'member_type', member_type, changed_list)
    # if changed_list:
    #     member.cost = cost
    #     # member.member_type = member_type
    #     member.save()
    #
    #     change_message = _('Changed %s.') % get_text_list(changed_list,
    #                                                       _('and')) if changed_list else ''
    #     prefix = u"【%s】" % batch.title
    #     LogEntry.objects.log_action(user_id=user.pk,
    #                                 content_type_id=ContentType.objects.get_for_model(member).pk,
    #                                 object_id=member.pk,
    #                                 object_repr=unicode(member),
    #                                 action_flag=CHANGE,
    #                                 change_message=(prefix + change_message) or _('No fields changed.'))
    #     msg = u"name: %s, %s" % (member.__unicode__(), u"情報が変更されました。")
    #     logger.info(msg)


def get_latest_concat(code):
    """最新の契約情報を取得する。

    :param code:
    :return:
    """
    concat = None
    if code:
        url = biz_config.get_config(constants.CONFIG_SERVICE_CONTRACT) % (code,)
        response = urllib2.urlopen(url)
        html = response.read()
        data = json.loads(html.replace("\r", "").replace("\n", ""))
        concat = data.get('contractDetail', None)
    return concat if concat else dict()


def get_cost(code):
    if code:
        date = '2017-03-01'
        url = 'http://service.e-business.co.jp:8080/ContractManagement/api/newContract?uid=%s' % (code,)
        response = urllib2.urlopen(url)
        html = response.read()
        data = json.loads(html.replace("\r", "").replace("\n", ""))
        period_list = []
        for item in data['contractList']:
            period_list.append(item['EMPLOYMENT_PERIOD_START'])
        latest_period = None
        if period_list:
            period_list.sort(reverse=True)
            for p in period_list:
                if date >= p:
                    latest_period = p
                    break
        for item in data['contractList']:
            if latest_period and item['EMPLOYMENT_PERIOD_START'] == latest_period:
                if item['ALLOWANLE_COST'] != "-":
                    return int(item['ALLOWANLE_COST']) if item['ALLOWANLE_COST'] else 0
        for item in data['contractList']:
            if item['EMPLOYER_NO'] == code:
                return int(item['ALLOWANLE_COST']) if item['ALLOWANLE_COST'] != "-" else 0
    return 0


def get_batch_manager(name):
    try:
        batch = models.BatchManage.objects.get(name=name)
    except ObjectDoesNotExist:
        batch = models.BatchManage(name=name)
    return batch


def get_members_information():
    all_members = models.get_on_sales_members()
    working_members = models.get_working_members()
    waiting_members = models.get_waiting_members()
    current_month_release = models.get_release_current_month()
    next_month_release = models.get_release_next_month()
    next_2_month_release = models.get_release_next_2_month()

    summary = {'all_member_count': all_members.count(),
               'working_member_count': working_members.count(),
               'waiting_member_count': waiting_members.count(),
               'current_month_count': current_month_release.count(),
               'next_month_count': next_month_release.count(),
               'next_2_month_count': next_2_month_release.count(),
               }

    status_list = []
    for salesperson in get_salesperson_members():
        d = dict()
        d['salesperson'] = salesperson
        d['all_member_count'] = salesperson.get_on_sales_members().count()
        d['working_member_count'] = salesperson.get_working_members().count()
        d['waiting_member_count'] = salesperson.get_waiting_members().count()
        d['current_month_count'] = salesperson.get_release_current_month().count()
        d['next_month_count'] = salesperson.get_release_next_month().count()
        d['next_2_month_count'] = salesperson.get_release_next_2_month().count()
        status_list.append(d)

    return status_list, summary


def notify_member_status_mails(batch, status_list, summary):
    """メールを通知する。

    :param batch バッチに管理ファイル
    :param status_list 通知の内容リスト
    :param summary 通知の集計情報
    """
    def get_status_info(salesperson_id):
        for status in status_list:
            if status['salesperson'].pk == salesperson_id:
                return [status]
        return []

    logger = batch.get_logger()
    today = datetime.date.today()
    next_month = common.add_months(today, 1)
    next_2_months = common.add_months(today, 2)
    next_ym = next_month.strftime('%Y%m')
    next_2_ym = next_2_months.strftime('%Y%m')
    # 営業部長取得する
    directors = get_salesperson_director()
    if directors:
        context = {'salesperson_list': directors,
                   'status_list': status_list,
                   'summary': summary,
                   'domain': biz_config.get_domain_name(),
                   'next_ym': next_ym,
                   'next_2_ym': next_2_ym,
                   }
        recipient_list = []
        for salesperson in directors:
            recipient_list.extend(salesperson.get_notify_mail_list())
        if recipient_list:
            batch.send_notify_mail(context, recipient_list)
    salesperson_list = get_salesperson_members()
    if salesperson_list:
        for salesperson in salesperson_list:
            recipient_list = salesperson.get_notify_mail_list()
            context = {'salesperson_list': [salesperson],
                       'status_list': get_status_info(salesperson.pk),
                       'summary': None,
                       'domain': biz_config.get_domain_name(),
                       'next_ym': next_ym,
                       'next_2_ym': next_2_ym,
                       }
            if recipient_list:
                batch.send_notify_mail(context, recipient_list, no_cc=True)
            else:
                logger.warning(u"%s の宛先が空白になっている。" % (salesperson.__unicode__(),))


def send_attendance_format(batch, date):
    """勤怠フォーマットを各部署の部長に送付する。

    :param date: 対象年月の出勤情報
    :param batch:
    :return:
    """
    logger = batch.get_logger()
    if not batch.attachment1 or not os.path.exists(batch.attachment1.path):
        logger.warning(u"出勤フォーマットの添付ファイルが設定していません。")
        return

    sections = biz.get_on_sales_section()
    if not date:
        date = datetime.datetime.today()
    for section in sections:
        statistician = section.get_attendance_statistician()
        if statistician.count() == 0:
            logger.warning(u"部署「%s」の勤務統計者が設定していません。" % (section.__unicode__(),))
            continue
        recipient_list = []
        name_list = []
        for member in statistician:
            recipient_list.extend(member.get_notify_mail_list())
            name_list.append(member.__unicode__())
        if not recipient_list:
            logger.error(u"%s の送信先メールアドレスが設定していません。" % (u",".join(name_list),))
            continue

        project_members = biz.get_project_members_month_section(section, date)
        output = file_gen.generate_attendance_format(batch.attachment1.path, project_members, date)

        context = {'statistician': statistician,
                   'section': section,
                   }
        attachment = (constants.NAME_SECTION_ATTENDANCE % (section.name, date.year, date.month) + ".xlsx",
                      output,
                      constants.MIME_TYPE_EXCEL)
        batch.send_notify_mail(context, recipient_list, [attachment])


def get_salesperson_director():
    """営業の管理者を取得する。
    """
    return models.Salesperson.objects.public_filter(member_type=0)


def get_salesperson_members():
    """営業のメンバーを取得する。
    """
    return models.Salesperson.objects.public_filter(member_type=5)


def members_to_excel(data_list, path):
    book = px.Workbook()
    sheet = book.create_sheet(title='Members')
    r = 2
    for i, data in enumerate(data_list):
        c = 1
        for key in data.keys():
            if i == 0:
                sheet.cell(row=1, column=c).value = key
            sheet.cell(row=r, column=c).value = data.get(key, None)
            c += 1
        r += 1

    book.save(path)
