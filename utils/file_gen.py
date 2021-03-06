# coding: UTF-8
"""
Created on 2015/10/29

@author: Yang Wanjun
"""
from __future__ import unicode_literals
import os
import re
import datetime
import StringIO
import xlsxwriter
import openpyxl as px
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font

try:
    import pythoncom
    import win32com.client
    is_win32 = True
except:
    is_win32 = False

import constants
import common
import errors
from eb import models


def generate_resume(member):
    output = StringIO.StringIO()

    book = xlsxwriter.Workbook(output)
    sheet = book.add_worksheet()
    # 全体の設定
    sheet.set_column('A:AP', 1.9)
    # タイトル設定
    title_format = book.add_format({'bold': True,
                                    'border': 1,
                                    'align': 'center',
                                    'valign': 'vcenter',
                                    'font_size': 24})
    sheet.merge_range('A1:AP3', u"技　術　者　経　歴　書", title_format)
    # 最終更新日
    title_format = book.add_format({'bold': True,
                                    'border': 0,
                                    'align': 'right',
                                    'valign': 'vcenter',
                                    'font_size': 11})
    content_format = book.add_format({'bold': True,
                                      'border': 0,
                                      'align': 'center',
                                      'valign': 'vcenter',
                                      'font_size': 11})
    sheet.merge_range('A4:AI4', u"最終更新日：", title_format)
    today = datetime.date.today()
    sheet.merge_range('AJ4:AP4', u"%d年%02d月" % (today.year, today.month), content_format)
    # 基本情報設定
    title_format_top = book.add_format({'bold': True,
                                        'top': 1,
                                        'left': 1,
                                        'right': 1,
                                        'align': 'center',
                                        'valign': 'vcenter',
                                        'bg_color': '#c0c0c0',
                                        'font_size': 10,
                                        'text_wrap': True})
    title_format_inner = book.add_format({'bold': True,
                                          'top': 4,
                                          'left': 1,
                                          'right': 1,
                                          'align': 'center',
                                          'valign': 'vcenter',
                                          'bg_color': '#c0c0c0',
                                          'font_size': 10,
                                          'text_wrap': True})
    title_format_bottom = book.add_format({'bold': True,
                                           'top': 4,
                                           'left': 1,
                                           'right': 1,
                                           'bottom': 1,
                                           'align': 'center',
                                           'valign': 'vcenter',
                                           'bg_color': '#c0c0c0',
                                           'font_size': 10,
                                           'text_wrap': True})
    content_format_top = book.add_format({'valign': 'vcenter',
                                          'font_size': 9,
                                          'right': 1,
                                          'top': 1})
    content_format_inner = book.add_format({'valign': 'vcenter',
                                            'font_size': 9,
                                            'right': 1,
                                            'top': 4})
    content_format_inner2 = book.add_format({'valign': 'vcenter',
                                             'font_size': 9,
                                             'top': 4})
    content_format_inner3 = book.add_format({'valign': 'vcenter',
                                             'font_size': 18,
                                             'top': 4})
    content_format_inner_center = book.add_format({'valign': 'vcenter',
                                                   'align': 'center',
                                                   'font_size': 9,
                                                   'right': 1,
                                                   'left': 4,
                                                   'top': 4})
    content_format_bottom = book.add_format({'valign': 'vcenter',
                                             'font_size': 9,
                                             'right': 1,
                                             'top': 4,
                                             'bottom': 1})
    sheet.merge_range('A5:C5', u"フリガナ", title_format_top)
    sheet.merge_range('A6:C7', u"氏　　名", title_format_inner)
    sheet.merge_range('A8:C9', u"現 住 所", title_format_inner)
    sheet.merge_range('A10:C10', u"最 寄 駅", title_format_inner)
    sheet.merge_range('A11:C12', u"在日年数", title_format_bottom)

    sheet.merge_range('D5:L5', member.first_name_ja + " " + member.last_name_ja if member.last_name_ja else "",
                      content_format_top)
    sheet.merge_range('D6:L7', member.get_resume_name(), content_format_inner3)
    sheet.merge_range('D8:L9', member.address1, content_format_inner)
    sheet.merge_range('D10:L10', member.nearest_station, content_format_inner)
    sheet.merge_range('D11:L12', u"%s 年" % (member.years_in_japan,) if member.years_in_japan else "",
                      content_format_bottom)

    sheet.merge_range('M5:O6', u"性　 　別", title_format_top)
    sheet.merge_range('M7:O8', u"本     籍", title_format_inner)
    sheet.merge_range('M9:O10', u"生年月日", title_format_inner)
    sheet.merge_range('M11:O12', u"婚　　姻\r\n状　　況", title_format_bottom)

    sheet.merge_range('P5:X6', member.get_sex_display(), content_format_top)
    sheet.merge_range('P7:X8', member.country, content_format_inner)
    date_format = u'%Y年%m月%d日'.encode('utf-8')
    sheet.merge_range('P9:T10', member.birthday.strftime(date_format).decode('utf-8') if member.birthday else "",
                      content_format_inner2)
    sheet.merge_range('U9:X10', str(member.get_age()), content_format_inner_center)
    sheet.merge_range('P11:X12', member.get_is_married_display(), content_format_bottom)

    sheet.merge_range('Y5:AA7', u"日 本 語\r\n／英語", title_format_top)
    sheet.merge_range('Y8:AA10', u"資　　格", title_format_inner)
    sheet.merge_range('Y11:AA12', u"得　　意", title_format_bottom)

    sheet.merge_range('AB5:AP7', member.japanese_description, content_format_top)
    sheet.merge_range('AB8:AP10', member.certificate, content_format_inner)
    sheet.merge_range('AB11:AP12', member.skill_description, content_format_bottom)

    # 学歴を書き込む
    title_format_1 = book.add_format({'bold': True,
                                      'align': 'center',
                                      'valign': 'vcenter',
                                      'bg_color': '#c0c0c0',
                                      'font_size': 18,
                                      'left': 1,
                                      'top': 1,
                                      'bottom': 1,
                                      'right': 4})
    title_format_2 = book.add_format({'bold': True,
                                      'align': 'center',
                                      'valign': 'vcenter',
                                      'bg_color': '#c0c0c0',
                                      'font_size': 10,
                                      'left': 4,
                                      'top': 1,
                                      'bottom': 4,
                                      'right': 4})
    title_format_3 = book.add_format({'bold': True,
                                      'align': 'center',
                                      'valign': 'vcenter',
                                      'bg_color': '#c0c0c0',
                                      'font_size': 10,
                                      'left': 4,
                                      'top': 1,
                                      'bottom': 4,
                                      'right': 1})
    content_format_inner = book.add_format({'font_size': 10,
                                            'bottom': 4,
                                            'left': 4,
                                            'right': 4})
    content_format_right = book.add_format({'font_size': 10,
                                            'bottom': 4,
                                            'left': 4,
                                            'right': 1})
    content_format_bottom = book.add_format({'font_size': 10,
                                             'bottom': 1,
                                             'left': 4,
                                             'right': 4})
    content_format_bottom_right = book.add_format({'font_size': 10,
                                                   'bottom': 1,
                                                   'left': 4,
                                                   'right': 1})
    sheet.merge_range('A14:E16', u"学    歴", title_format_1)
    sheet.merge_range('F14:N14', u"期       間", title_format_2)
    sheet.merge_range('O14:AP14', u"学校名称　／　学部　／　専門　／学位", title_format_3)
    for i, degree in enumerate(member.degree_set.all()):
        period = u"%d/%02d/%02d - %d/%02d/%02d" % (degree.start_date.year, degree.start_date.month,
                                                   degree.start_date.day,
                                                   degree.end_date.year, degree.end_date.month, degree.end_date.day)
        sheet.merge_range(14 + i, 5, 14 + i, 13, period,
                          content_format_inner if i == 0 else content_format_bottom)
        sheet.merge_range(14 + i, 14, 14 + i, 41, degree.description,
                          content_format_right if i == 0 else content_format_bottom_right)
    else:
        cnt = member.degree_set.all().count()
        if cnt == 1:
            sheet.merge_range(14 + 1, 5, 14 + 1, 13, "", content_format_bottom)
            sheet.merge_range(14 + 1, 14, 14 + 1, 41, "", content_format_bottom_right)
        elif cnt == 0:
            sheet.merge_range(14 + 0, 5, 14 + 0, 13, "", content_format_inner)
            sheet.merge_range(14 + 0, 14, 14 + 0, 41, "", content_format_right)
            sheet.merge_range(14 + 1, 5, 14 + 1, 13, "", content_format_bottom)
            sheet.merge_range(14 + 1, 14, 14 + 1, 41, "", content_format_bottom_right)

    # 業務経歴を書き込む
    title_format = book.add_format({'bold': True,
                                    'align': 'center',
                                    'font_size': 18,
                                    'bg_color': '#ffff99',
                                    'border': 1})
    title_format_left_top = book.add_format({'bold': True,
                                             'align': 'center',
                                             'valign': 'vcenter',
                                             'bg_color': '#c0c0c0',
                                             'font_size': 10,
                                             'top': 1,
                                             'left': 1,
                                             'bottom': 1,
                                             'right': 4})
    title_format_inner = book.add_format({'bold': True,
                                          'align': 'center',
                                          'valign': 'vcenter',
                                          'bg_color': '#c0c0c0',
                                          'font_size': 10,
                                          'top': 1,
                                          'left': 4,
                                          'bottom': 1,
                                          'right': 4,
                                          'text_wrap': True})
    title_format_right_top = book.add_format({'bold': True,
                                              'align': 'center',
                                              'valign': 'vcenter',
                                              'bg_color': '#c0c0c0',
                                              'font_size': 10,
                                              'top': 1,
                                              'left': 4,
                                              'bottom': 4,
                                              'right': 1,
                                              'text_wrap': True})
    title_format_bottom = book.add_format({'bold': True,
                                           'align': 'center',
                                           'valign': 'vcenter',
                                           'bg_color': '#c0c0c0',
                                           'font_size': 10,
                                           'top': 4,
                                           'left': 4,
                                           'bottom': 1,
                                           'right': 4,
                                           'text_wrap': True})
    title_format_bottom_right = book.add_format({'bold': True,
                                                 'align': 'center',
                                                 'valign': 'vcenter',
                                                 'bg_color': '#c0c0c0',
                                                 'font_size': 10,
                                                 'top': 4,
                                                 'left': 4,
                                                 'bottom': 1,
                                                 'right': 1,
                                                 'text_wrap': True})
    content_format_left = book.add_format({'align': 'center',
                                           'valign': 'vcenter',
                                           'font_size': 9,
                                           'left': 1,
                                           'top': 1,
                                           'bottom': 1,
                                           'right': 4})
    content_format_inner = book.add_format({'valign': 'vcenter',
                                            'font_size': 9,
                                            'left': 4,
                                            'top': 1,
                                            'bottom': 1,
                                            'right': 4,
                                            'text_wrap': True})
    content_format_inner2 = book.add_format({'valign': 'vcenter',
                                             'font_size': 9,
                                             'left': 4,
                                             'top': 1,
                                             'bottom': 4,
                                             'right': 4})
    content_format_inner3 = book.add_format({'align': 'center',
                                             'valign': 'vcenter',
                                             'font_size': 9,
                                             'left': 4,
                                             'top': 1,
                                             'bottom': 4,
                                             'right': 4})
    content_format_inner4 = book.add_format({'valign': 'vcenter',
                                            'font_size': 9,
                                            'left': 4,
                                            'top': 4,
                                            'bottom': 1,
                                            'right': 4,
                                            'text_wrap': True})
    content_format_right = book.add_format({'valign': 'vcenter',
                                            'font_size': 9,
                                            'left': 4,
                                            'top': 1,
                                            'bottom': 1,
                                            'right': 1,
                                            'text_wrap': True})
    sheet.merge_range('A18:AP19', u"業  務  経  歴", title_format)
    sheet.merge_range('A20:A24', u"No.", title_format_left_top)
    sheet.merge_range('B20:E24', u"作業期間", title_format_inner)
    sheet.merge_range('F20:Q24', u"業務内容", title_format_inner)
    sheet.merge_range('R20:W24', u"機種／OS", title_format_inner)
    sheet.merge_range('X20:AC24', u"言語／ツール\r\nＤＢ", title_format_inner)
    sheet.merge_range('AD20:AF24', u"作\r\n業\r\n区\r\n分", title_format_inner)
    sheet.merge_range('AG20:AP20', u"作業工程", title_format_right_top)
    sheet.merge_range('AG21:AG24', u"要\r\n件\r\n定\r\n義", title_format_bottom)
    sheet.merge_range('AH21:AH24', u"調\r\n査\r\n分\r\n析", title_format_bottom)
    sheet.merge_range('AI21:AI24', u"基\r\n本\r\n設\r\n計", title_format_bottom)
    sheet.merge_range('AJ21:AJ24', u"詳\r\n細\r\n設\r\n計", title_format_bottom)
    sheet.merge_range('AK21:AK24', u"開\r\n発\r\n製\r\n造", title_format_bottom)
    sheet.merge_range('AL21:AL24', u"単\r\n体\r\n試\r\n験", title_format_bottom)
    sheet.merge_range('AM21:AM24', u"結\r\n合\r\n試\r\n験", title_format_bottom)
    sheet.merge_range('AN21:AN24', u"総\r\n合\r\n試\r\n験", title_format_bottom)
    sheet.merge_range('AO21:AO24', u"保\r\n守\r\n運\r\n用", title_format_bottom)
    sheet.merge_range('AP21:AP24', u"サ\r\nポ\r\nー\r\nト", title_format_bottom_right)

    project_count = member.projectmember_set.all().count()
    all_project_count = member.historyproject_set.all().count() + project_count

    # 案件一覧
    for i, project_member in enumerate(member.projectmember_set.all()):
        first_row = 24 + (i * 3)
        sheet.set_row(first_row + 1, 30)
        sheet.set_row(first_row + 2, 30)
        sheet.merge_range(first_row, 0, first_row + 2, 0, str(i + 1), content_format_left)
        period = u"%d年%02d月～\r\n%d年%02d月" % (project_member.project.start_date.year,
                                                 project_member.project.start_date.month,
                                                 project_member.project.end_date.year,
                                                 project_member.project.end_date.month)
        sheet.merge_range(first_row, 1, first_row + 2, 4, period, content_format_inner)
        sheet.merge_range(first_row, 5, first_row, 13, project_member.project.name, content_format_inner2)
        sheet.merge_range(first_row, 14, first_row, 16, "", content_format_inner3)
        sheet.merge_range(first_row + 1, 5, first_row + 2, 16, project_member.project.description,
                          content_format_inner4)
        os_list = [os.name for os in project_member.project.os.all()]
        sheet.merge_range(first_row, 17, first_row + 2, 22, u"\r\n".join(os_list), content_format_inner)
        skill_list = [skill.name for skill in project_member.project.skills.all()]
        sheet.merge_range(first_row, 23, first_row + 2, 28, u"\r\n".join(skill_list), content_format_inner)
        sheet.merge_range(first_row, 29, first_row + 2, 31, project_member.role, content_format_inner)
        sheet.merge_range(first_row, 32, first_row + 2, 32, u"●" if project_member.is_in_rd() else "",
                          content_format_inner)
        sheet.merge_range(first_row, 33, first_row + 2, 33, u"●" if project_member.is_in_sa() else "",
                          content_format_inner)
        sheet.merge_range(first_row, 34, first_row + 2, 34, u"●" if project_member.is_in_bd() else "",
                          content_format_inner)
        sheet.merge_range(first_row, 35, first_row + 2, 35, u"●" if project_member.is_in_dd() else "",
                          content_format_inner)
        sheet.merge_range(first_row, 36, first_row + 2, 36, u"●" if project_member.is_in_pg() else "",
                          content_format_inner)
        sheet.merge_range(first_row, 37, first_row + 2, 37, u"●" if project_member.is_in_pt() else "",
                          content_format_inner)
        sheet.merge_range(first_row, 38, first_row + 2, 38, u"●" if project_member.is_in_it() else "",
                          content_format_inner)
        sheet.merge_range(first_row, 39, first_row + 2, 39, u"●" if project_member.is_in_st() else "",
                          content_format_inner)
        sheet.merge_range(first_row, 40, first_row + 2, 40, u"●" if project_member.is_in_maintain() else "",
                          content_format_inner)
        sheet.merge_range(first_row, 41, first_row + 2, 41, u"●" if project_member.is_in_support() else "",
                          content_format_right)

    # 以前やっていた案件を出力する。
    for i, project in enumerate(member.historyproject_set.all()):
        first_row = 24 + (project_count * 3) + (i * 3)
        sheet.set_row(first_row + 1, 30)
        sheet.set_row(first_row + 2, 30)
        sheet.merge_range(first_row, 0, first_row + 2, 0, str(project_count + i + 1), content_format_left)
        period = u"%d年%02d月～\r\n%d年%02d月" % (project.start_date.year, project.start_date.month,
                                             project.end_date.year, project.end_date.month)
        sheet.merge_range(first_row, 1, first_row + 2, 4, period, content_format_inner)
        sheet.merge_range(first_row, 5, first_row, 13, project.name, content_format_inner2)
        sheet.merge_range(first_row, 14, first_row, 16, project.location, content_format_inner3)
        sheet.merge_range(first_row + 1, 5, first_row + 2, 16, project.description, content_format_inner4)
        os_list = [os.name for os in project.os.all()]
        sheet.merge_range(first_row, 17, first_row + 2, 22, u"\r\n".join(os_list), content_format_inner)
        skill_list = [skill.name for skill in project.skill.all()]
        sheet.merge_range(first_row, 23, first_row + 2, 28, u"\r\n".join(skill_list), content_format_inner)
        sheet.merge_range(first_row, 29, first_row + 2, 31, project.role, content_format_inner)
        sheet.merge_range(first_row, 32, first_row + 2, 32, u"●" if project.is_in_rd() else "", content_format_inner)
        sheet.merge_range(first_row, 33, first_row + 2, 33, u"●" if project.is_in_sa() else "", content_format_inner)
        sheet.merge_range(first_row, 34, first_row + 2, 34, u"●" if project.is_in_bd() else "", content_format_inner)
        sheet.merge_range(first_row, 35, first_row + 2, 35, u"●" if project.is_in_dd() else "", content_format_inner)
        sheet.merge_range(first_row, 36, first_row + 2, 36, u"●" if project.is_in_pg() else "", content_format_inner)
        sheet.merge_range(first_row, 37, first_row + 2, 37, u"●" if project.is_in_pt() else "", content_format_inner)
        sheet.merge_range(first_row, 38, first_row + 2, 38, u"●" if project.is_in_it() else "", content_format_inner)
        sheet.merge_range(first_row, 39, first_row + 2, 39, u"●" if project.is_in_st() else "", content_format_inner)
        sheet.merge_range(first_row, 40, first_row + 2, 40, u"●" if project.is_in_maintain() else "",
                          content_format_inner)
        sheet.merge_range(first_row, 41, first_row + 2, 41, u"●" if project.is_in_support() else "",
                          content_format_right)
    if all_project_count == 0:
        all_project_count = 13
        for i in range(13):
            first_row = 24 + (i * 3)
            sheet.merge_range(first_row, 0, first_row + 2, 0, str(i + 1), content_format_left)
            sheet.merge_range(first_row, 1, first_row + 2, 4, "", content_format_inner)
            sheet.merge_range(first_row, 5, first_row, 13, "", content_format_inner2)
            sheet.merge_range(first_row, 14, first_row, 16, "", content_format_inner3)
            sheet.merge_range(first_row + 1, 5, first_row + 2, 16, "", content_format_inner4)
            sheet.merge_range(first_row, 17, first_row + 2, 22, "", content_format_inner)
            sheet.merge_range(first_row, 23, first_row + 2, 28, "", content_format_inner)
            sheet.merge_range(first_row, 29, first_row + 2, 31, "", content_format_inner)
            sheet.merge_range(first_row, 32, first_row + 2, 32, "", content_format_inner)
            sheet.merge_range(first_row, 33, first_row + 2, 33, "", content_format_inner)
            sheet.merge_range(first_row, 34, first_row + 2, 34, "", content_format_inner)
            sheet.merge_range(first_row, 35, first_row + 2, 35, "", content_format_inner)
            sheet.merge_range(first_row, 36, first_row + 2, 36, "", content_format_inner)
            sheet.merge_range(first_row, 37, first_row + 2, 37, "", content_format_inner)
            sheet.merge_range(first_row, 38, first_row + 2, 38, "", content_format_inner)
            sheet.merge_range(first_row, 39, first_row + 2, 39, "", content_format_inner)
            sheet.merge_range(first_row, 40, first_row + 2, 40, "", content_format_inner)
            sheet.merge_range(first_row, 41, first_row + 2, 41, "", content_format_right)

    row_tail = 24 + (all_project_count * 3)
    tail_format = book.add_format({'align': 'center',
                                   'valign': 'vcenter',
                                   'font_size': 9,
                                   'border': 1})
    sheet.merge_range(row_tail, 0, row_tail, 41,
                      u"作業区分：   M：ﾏﾈｰｼﾞｬｰ、L：ﾘｰﾀﾞｰ、SL：ｻﾌﾞﾘｰﾀﾞｰ、SE：.ｼｽﾃﾑｴﾝｼﾞﾆｱ、SP：ｼｽﾃﾑﾌﾟﾛｸﾞﾗﾏｰ、PG：ﾌﾟﾛｸﾞﾗﾏｰ、OP：ｵﾍﾟﾚｰﾀｰ",
                      tail_format)

    book.close()
    output.seek(0)
    return output


def generate_quotation(project, user, company):
    """見積書を生成する。

    Arguments：
      project: 対象の案件
      company: 会社名

    Returns：
      dict

    Raises：
      FileNotExistException
    """
    pythoncom.CoInitializeEx(pythoncom.COINIT_MULTITHREADED)
    template_file = project.client.quotation_file if project.client.quotation_file else company.quotation_file
    if not template_file or not os.path.exists(template_file.path):
        raise errors.FileNotExistException(constants.ERROR_TEMPLATE_NOT_EXISTS)

    template_book = get_excel_template(template_file.path)
    template_sheet = template_book.Worksheets(1)
    book = get_new_book()
    cnt = book.Sheets.Count
    # テンプレートを生成対象ワークブックにコピーする。
    template_sheet.Copy(None, book.Worksheets(cnt))
    template_book.Close()
    sheet = book.Worksheets(cnt + 1)

    data = dict()
    # お客様会社名
    data['CLIENT_COMPANY_NAME'] = project.client.name
    # 見積書番号
    data['QUOTATION_NO'] = common.get_quotation_no(user)
    # 案件名称
    data['PROJECT_NAME'] = project.name
    # 業務内容
    data['PROJECT_DESCRIPTION'] = project.description
    # 開発期間
    data['START_DATE'] = u"%s年%s月%s日" % (project.start_date.year, project.start_date.month, project.start_date.day)
    data['END_DATE'] = u"%s年%s月%s日" % (project.end_date.year, project.end_date.month, project.end_date.day)
    # 基準時間
    data['MIN_HOUR'] = u"%d" % (project.min_hours,)
    data['MAX_HOUR'] = u"%d" % (project.max_hours,)
    # 開発場所
    data['ADDRESS'] = project.address

    # メンバー毎の料金
    detail_members = []
    for project_member in project.projectmember_set.public_all():
        dict_member = dict()
        dict_member['ITEM_NAME'] = project_member.member.__unicode__()
        dict_member['ITEM_PRICE'] = u"￥%s人/月" % (project_member.price,)
        dict_member['ITEM_PLUS'] = project_member.plus_per_hour
        dict_member['ITEM_MINUS'] = project_member.minus_per_hour
        detail_members.append(dict_member)

    data['quotation_details'] = detail_members
    data['details_start_col'] = 3

    replace_excel_dict(sheet, data)

    for i in range(cnt, 0, -1):
        book.Worksheets(i).Delete()

    file_folder = os.path.join(os.path.dirname(template_file.path), "temp")
    if not os.path.exists(file_folder):
        os.mkdir(file_folder)
    file_name = "tmp_%s_%s.xls" % (constants.DOWNLOAD_QUOTATION, datetime.datetime.now().strftime("%Y%m%d_%H%M%S%f"))
    path = os.path.join(file_folder, file_name)
    book.SaveAs(path, FileFormat=constants.EXCEL_FORMAT_EXCEL2003)
    return path


def generate_request_linux(project, data, request_no, ym):
    if isinstance(project, models.Project):
        path = common.get_request_file_path(request_no, project.client.name, ym)
        is_lump = project.is_lump
    else:
        path = common.get_subcontractor_request_file_path(request_no, unicode(project), ym)
        is_lump = False
    book = xlsxwriter.Workbook(path)
    sheet = book.add_worksheet()

    # タイトル設定
    title_format = book.add_format({'bold': True,
                                    'align': 'center',
                                    'valign': 'vcenter',
                                    'font_size': 18,
                                    'underline': 2})
    sheet.merge_range('A1:P1', u"　　御　請　求　書　　", title_format)
    sheet.write_string('M3', u"請求番号")
    sheet.write_number('O3', int(request_no))
    sheet.write_string('M4', u"発  行 日")
    sheet.write_string('O4', data['DETAIL']['PUBLISH_DATE'])
    sheet.write_string('B3', u"〒" + data['DETAIL']['CLIENT_POST_CODE'])
    sheet.write_string('B4', data['DETAIL']['CLIENT_ADDRESS'])
    sheet.write_string('B6', u"Tel: " + data['DETAIL']['CLIENT_TEL'])
    name_format = book.add_format({'bold': True,
                                   'font_size': 12,
                                   'underline': 1})
    sheet.write_string('B8', data['DETAIL']['CLIENT_COMPANY_NAME'] + u"御中", name_format)
    sheet.write_string('B10', u"　下記のとおりご請求申し上げます。")
    format1 = book.add_format({'bold': True,
                               'font_size': 12})
    sheet.write_string('B12', u"御請求額　 ：　", format1)
    format1 = book.add_format({'bold': True,
                               'font_size': 14,
                               'underline': 2})
    sheet.write_string('E12', u"\\" + data['DETAIL']['ITEM_AMOUNT_ALL_COMMA'] + u"円", format1)
    sheet.write_string('B14', u"作業期間　   ：")
    sheet.write_string('E14', data['DETAIL']['WORK_PERIOD'])
    sheet.write_string('B16', u"注文番号　   ：")
    sheet.write_string('E16', data['DETAIL']['ORDER_NO'])
    sheet.write_string('B18', u"注文日　　　  ：")
    sheet.write_string('E18', data['DETAIL']['REQUEST_DATE'])
    sheet.write_string('B20', u"契約件名　　 ：　")
    sheet.write_string('E20', data['DETAIL']['CONTRACT_NAME'])
    sheet.write_string('B22', u"お支払い期限　：")
    sheet.write_string('E22', data['DETAIL']['REMIT_DATE'])
    sheet.write_string('M10', u"〒" + data['DETAIL']['POST_CODE'])
    sheet.write_string('M11', data['DETAIL']['ADDRESS'])
    sheet.write_string('M12', data['DETAIL']['COMPANY_NAME'])
    sheet.write_string('M13', u"代表取締役　　" + data['DETAIL']['MASTER'])
    sheet.write_string('M14', u"TEL：" + data['DETAIL']['TEL'])
    sheet.insert_textbox('M15', '', {'width': 90,
                                     'height': 90,
                                     'x_offset': 4,
                                     'y_offset': 6,
                                     'align': {'vertical': 'middle', 'horizontal': 'center'}
                                     }
                         )
    sheet.insert_textbox('O15', '', {'width': 90,
                                     'height': 90,
                                     'x_offset': 6,
                                     'y_offset': 6,
                                     'align': {'vertical': 'middle', 'horizontal': 'center'}
                                     }
                         )
    sheet.insert_textbox('P15', '', {'width': 90,
                                     'height': 90,
                                     'x_offset': -2,
                                     'y_offset': 6,
                                     'align': {'vertical': 'middle', 'horizontal': 'center'}
                                     }
                         )
    title_format = book.add_format({'font_size': 11,
                                    'border': 1,
                                    'align': 'center',
                                    'valign': 'vcenter', })
    cell_format = book.add_format({'font_size': 11,
                                   'border': 1, })
    range1_format = book.add_format({'font_size': 11,
                                    'left': 1,
                                    'top': 1,
                                    'bottom': 1})
    range2_format = book.add_format({'font_size': 11,
                                    'top': 1,
                                    'bottom': 1})
    range3_format = book.add_format({'font_size': 11,
                                    'right': 1,
                                    'top': 1,
                                    'bottom': 1})
    num_format = book.add_format({'num_format': '#,###', 'border': 1})
    float_format = book.add_format({'num_format': '#,###.00', 'border': 1})
    start_row = 24

    def border_row(row_index, is_lump_temp=False):
        sheet.write_string(row_index, 1, '', cell_format)
        sheet.write_string(row_index, 14, '', cell_format)
        sheet.write_string(row_index, 15, '', cell_format)
        if is_lump_temp:
            sheet.write_string(row_index, 2, '', range1_format)
            sheet.write_string(row_index, 3, '', range2_format)
            sheet.write_string(row_index, 4, '', range2_format)
            sheet.write_string(row_index, 5, '', range2_format)
            sheet.write_string(row_index, 6, '', range2_format)
            sheet.write_string(row_index, 7, '', range2_format)
            sheet.write_string(row_index, 8, '', range2_format)
            sheet.write_string(row_index, 9, '', range2_format)
            sheet.write_string(row_index, 10, '', range3_format)
            sheet.write_string(row_index, 11, '', range1_format)
            sheet.write_string(row_index, 12, '', range2_format)
            sheet.write_string(row_index, 13, '', range3_format)
        else:
            sheet.write_string(row_index, 2, '', range1_format)
            sheet.write_string(row_index, 3, '', range2_format)
            sheet.write_string(row_index, 4, '', range2_format)
            sheet.write_string(row_index, 5, '', range2_format)
            sheet.write_string(row_index, 6, '', range3_format)
            sheet.write_string(row_index, 7, '', cell_format)
            sheet.write_string(row_index, 8, '', cell_format)
            sheet.write_string(row_index, 9, '', cell_format)
            sheet.write_string(row_index, 10, '', cell_format)
            sheet.write_string(row_index, 11, '', cell_format)
            sheet.write_string(row_index, 12, '', cell_format)
            sheet.write_string(row_index, 13, '', cell_format)

    if data['MEMBERS']:
        sheet.write_string('B24', u"番号", title_format)
        sheet.merge_range('C24:G24', u"項　　　　目", title_format)
        sheet.write_string('H24', u"単価", title_format)
        sheet.write_string('I24', u"作業H", title_format)
        sheet.write_string('J24', u"率", title_format)
        sheet.write_string('K24', u"Min/MaxH", title_format)
        sheet.write_string('L24', u"減", title_format)
        sheet.write_string('M24', u"増", title_format)
        sheet.write_string('N24', u"その他", title_format)
        sheet.write_string('O24', u"金額", title_format)
        sheet.write_string('P24', u"備考", title_format)
        for item in data['MEMBERS']:
            sheet.write_number(start_row, 1, int(item['NO']), cell_format)
            sheet.merge_range(start_row, 2, start_row, 6, item['ITEM_NAME'], cell_format)
            sheet.write_number(start_row, 7, item['ITEM_PRICE'], num_format)
            if item['ITEM_WORK_HOURS']:
                sheet.write_number(start_row, 8, float(item['ITEM_WORK_HOURS']), float_format)
            else:
                sheet.write_string(start_row, 8, '', cell_format)
            sheet.write_number(start_row, 9, item['ITEM_RATE'], float_format)
            sheet.write_string(start_row, 10, item['ITEM_MIN_MAX'], cell_format)
            sheet.write_number(start_row, 11, item['ITEM_MINUS_PER_HOUR'], num_format)
            sheet.write_number(start_row, 12, item['ITEM_PLUS_PER_HOUR'], num_format)
            sheet.write_string(start_row, 13, item['ITEM_OTHER'], cell_format)
            sheet.write_number(start_row, 14, item['ITEM_AMOUNT_TOTAL'], num_format)
            sheet.write_string(start_row, 15, item['ITEM_COMMENT'], cell_format)
            start_row += 1
    elif data['detail_all']:
        item = data['detail_all']
        sheet.write_string('B24', u"番号", title_format)
        sheet.merge_range('C24:K24', u"項　　　　目", title_format)
        sheet.merge_range('L24:N24', u"単位", title_format)
        sheet.write_string('O24', u"金額", title_format)
        sheet.write_string('P24', u"備考", title_format)
        sheet.write_number('B25', int(item['NO']), cell_format)
        sheet.merge_range('C25:K25', item['ITEM_NAME_ATTENDANCE_TOTAL'], cell_format)
        sheet.merge_range('L25:N25', item['ITEM_UNIT'], cell_format)
        sheet.write_number('O25', data['DETAIL']['ITEM_AMOUNT_ATTENDANCE'], num_format)
        sheet.write_string('P25', item['ITEM_COMMENT'], cell_format)
        start_row += 1
    if start_row < 44:
        for i in range(start_row, 44):
            border_row(i, is_lump)
        start_row = 44
    else:
        start_row += 1
    for i in range(start_row, start_row + 5):
        border_row(i, is_lump)
    sheet.merge_range(start_row + 0, 3, start_row + 0, 5, u"（小計）", range2_format)
    sheet.write_number(start_row + 0, 14, data['DETAIL']['ITEM_AMOUNT_ATTENDANCE'], num_format)
    sheet.merge_range(start_row + 1, 3, start_row + 1, 5, u"(消費税）", range2_format)
    sheet.write_number(start_row + 1, 14, data['DETAIL']['ITEM_AMOUNT_ATTENDANCE_TAX'], num_format)
    sheet.merge_range(start_row + 2, 3, start_row + 2, 5, u"(合計）", range2_format)
    sheet.write_number(start_row + 2, 14, data['DETAIL']['ITEM_AMOUNT_ATTENDANCE_ALL'], num_format)
    sheet.merge_range(start_row + 3, 3, start_row + 3, 5, u"[控除、追加]", range2_format)
    sheet.write_string(start_row + 4, 1, u"控除", cell_format)
    start_row += 5
    if data['EXPENSES']:
        for i, item in enumerate(data['EXPENSES']):
            border_row(start_row, is_lump)
            if i == 0:
                sheet.write_string(start_row, 1, u"追加", cell_format)
            # sheet.write_string(start_row, 3, item['ITEM_EXPENSES_CATEGORY_SUMMARY'], range2_format)
            sheet.merge_range(start_row, 3, start_row, 13, item['ITEM_EXPENSES_CATEGORY_SUMMARY'], range3_format)
            sheet.write_number(start_row, 14, item['ITEM_EXPENSES_CATEGORY_AMOUNT'], num_format)
            start_row += 1
    else:
        border_row(start_row, is_lump)
        sheet.write_string(start_row, 1, u"追加", cell_format)
        start_row += 1
    border_row(start_row, is_lump)
    sheet.merge_range(start_row, 3, start_row, 4, u"(総計）", range2_format)
    sheet.write_number(start_row, 14, data['DETAIL']['ITEM_AMOUNT_ALL'], num_format)
    sheet.write_string(start_row + 1, 1, u"お振込銀行口座")
    sheet.write_string(start_row + 2, 2, data['DETAIL']['BANK_NAME'])
    sheet.write_string(start_row + 3, 2, u"%s（%s）" % (data['DETAIL']['BRANCH_NAME'], data['DETAIL']['BRANCH_NO']))
    sheet.write_string(start_row + 4, 2, u"%s　%s" % (data['DETAIL']['ACCOUNT_TYPE'], data['DETAIL']['ACCOUNT_NUMBER']))
    sheet.write_string(start_row + 5, 2, u"名義　　　　%s" % (data['DETAIL']['BANK_ACCOUNT_HOLDER'],))
    border_right_format = book.add_format({'right': 1})
    border_top_format = book.add_format({'top': 1})
    for i in range(start_row + 1, start_row + 6):
        sheet.write_string(i, 0, '', border_right_format)
    for i in range(start_row + 1, start_row + 6):
        sheet.write_string(i, 15, '', border_right_format)
    for i in range(1, 16):
        sheet.write_string(start_row + 6, i, '', border_top_format)

    # 全体の設定
    sheet.hide_gridlines()
    sheet.fit_to_pages(1, 1)
    sheet.set_column('A:A', 0.9)
    sheet.set_column('B:B', 4.9)
    sheet.set_column('C:C', 2.9)
    sheet.set_column('D:D', 2.9)
    sheet.set_column('E:E', 3.0)
    sheet.set_column('F:F', 3.0)
    sheet.set_column('G:G', 3.0)
    sheet.set_column('H:H', 10.3)
    sheet.set_column('I:I', 10.7)
    sheet.set_column('J:J', 4.9)
    sheet.set_column('K:K', 15.3)
    sheet.set_column('L:L', 4.9)
    sheet.set_column('M:M', 4.9)
    sheet.set_column('N:N', 6.6)
    sheet.set_column('O:O', 13.6)
    sheet.set_column('P:P', 12.4)
    sheet.set_row(0, 23.25)
    for i in range(1, 23):
        sheet.set_row(i, 15.5)
    book.close()
    return path


def generate_request(company, project, data, request_no, ym):
    """請求書を生成する。

    :param company:
    :param project:
    :param data:
    :param request_no: 請求番号
    :param ym:
    :return:
    """
    if is_win32:
        pythoncom.CoInitializeEx(pythoncom.COINIT_MULTITHREADED)
        if project.is_lump:
            request_file = company.request_lump_file
        else:
            request_file = project.client.request_file if project.client.request_file else company.request_file
        if not request_file or not os.path.exists(request_file.path):
            raise errors.FileNotExistException(constants.ERROR_TEMPLATE_NOT_EXISTS)

        template_book = get_excel_template(request_file.path)
        template_sheet = template_book.Worksheets(1)
        book = get_new_book()
        cnt = book.Sheets.Count
        # テンプレートを生成対象ワークブックにコピーする。
        template_sheet.Copy(None, book.Worksheets(cnt))
        template_book.Close()
        sheet = book.Worksheets(cnt + 1)

        replace_excel_dict(sheet, data['DETAIL'])
        replace_excel_dict(sheet, data['detail_all'])
        if not project.is_lump:
            # 一括案件の場合はアサインしたメンバーは表示しないように
            replace_excel_list(sheet, data['MEMBERS'])
            replace_excel_list(sheet, data['EXPENSES'], range_start="EXPENSES_START", range_end="EXPENSES_END")

        for i in range(cnt, 0, -1):
            book.Worksheets(i).Delete()

        path = common.get_request_file_path(request_no, project.client.name, ym)
        book.SaveAs(path)

        return path
    else:
        return generate_request_linux(project, data, request_no, ym)


def get_excel_template(path_file):
    if not os.path.exists(path_file):
        raise errors.FileNotExistException(constants.ERROR_TEMPLATE_NOT_EXISTS)

    xl_app = win32com.client.dynamic.Dispatch(constants.EXCEL_APPLICATION)
    xl_app.DisplayAlerts = False
    xl_app.Visible = 0
    book = xl_app.Workbooks.Open(path_file)
    return book


def get_new_book():
    xl_app = win32com.client.dynamic.Dispatch(constants.EXCEL_APPLICATION)
    xl_app.DisplayAlerts = False
    xl_app.Visible = 0
    book = xl_app.Workbooks.Add()
    return book


def find_cell_by_string(sheet, s, after=None):
    if after:
        return sheet.Cells.Find(What=s, LookIn=constants.xlFormulas, After=after, LookAt=constants.xlPart,
                                SearchOrder=constants.xlByRows, SearchDirection=constants.xlNext,
                                MatchCase=False, MatchByte=False, SearchFormat=False)
    else:
        return sheet.Cells.Find(What=s, LookIn=constants.xlFormulas, LookAt=constants.xlPart,
                                SearchOrder=constants.xlByRows, SearchDirection=constants.xlNext,
                                MatchCase=False, MatchByte=False, SearchFormat=False)


def get_replace_positions(sheet, data):
    """各置換文字列の位置を取得する。

    Arguments：
      sheet: エクセルのシート
      data: 置換する対象

    Returns：
      なし

    Raises：
      なし
    """
    d = dict()
    for key in data.keys():
        replace_key = "{$%s$}" % (key,)
        cell = find_cell_by_string(sheet, replace_key)
        if cell:
            d[key] = (cell.Row, cell.Column)
    return d


def get_replace_labels_position(sheet, start_row, start_col, row_span):
    labels = []
    for row in range(start_row, start_row + row_span):
        for col in range(start_col, sheet.UsedRange.Columns.Count):
            val = sheet.Cells(row, col).Value
            if val and val.strip() and val.find("{$") < 0:
                labels.append((row, col, val))
    return labels


def generate_order(data, template_path, is_request=False):
    """註文書を生成する。

    :param data 註文書の出力データ
    :param template_path: テンプレートのパス
    :param is_request: 注文請書
    :return エクセルのバイナリー
    """
    if False:
        # テンプレートを取得する
        if not template_path or not os.path.exists(template_path):
            raise errors.FileNotExistException(constants.ERROR_TEMPLATE_NOT_EXISTS)

        pythoncom.CoInitializeEx(pythoncom.COINIT_MULTITHREADED)
        # 新しいエクセルを作成する。
        template_book = get_excel_template(template_path)
        template_sheet = template_book.Worksheets(1)
        book = get_new_book()
        cnt = book.Sheets.Count
        # テンプレートを生成対象ワークブックにコピーする。
        template_sheet.Copy(None, book.Worksheets(cnt))
        template_book.Close()
        sheet = book.Worksheets(cnt + 1)
        replace_excel_dict(sheet, data['DETAIL'])
        replace_excel_list(sheet, data['MEMBERS'])

        for i in range(cnt, 0, -1):
            book.Worksheets(i).Delete()

        file_folder = os.path.join(os.path.dirname(template_path), "temp")
        if not os.path.exists(file_folder):
            os.mkdir(file_folder)
        file_name = "tmp_%s_%s.xls" % (constants.DOWNLOAD_ORDER, datetime.datetime.now().strftime("%Y%m%d_%H%M%S%f"))
        path = os.path.join(file_folder, file_name)
        # 一時ファイルを削除する。
        common.delete_temp_files(os.path.dirname(path))
        book.SaveAs(path, FileFormat=constants.EXCEL_FORMAT_EXCEL2003)

        return path
    else:
        return generate_order_linux(data, template_path, is_request)


def generate_order_linux(data, template_path, is_request):
    """openpyxlを利用してＢＰ注文書を作成する。

    :param data:
    :param template_path:
    :param is_request: 注文請書
    :return:
    """
    book = px.load_workbook(template_path)
    sheet = book.get_sheet_by_name('Sheet1')

    order_no = data['DETAIL']['ORDER_NO']
    partner_name = data['DETAIL']['SUBCONTRACTOR_NAME']
    path = common.get_order_file_path(order_no, partner_name, data['DETAIL']['YM'], is_request)

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, basestring):
                replaced_string = get_openpyxl_replaced_string(cell.value, data['DETAIL'])
                if replaced_string is not None:
                    if cell.value.find('{$ORDER_NO$}') >= 0 or cell.value.find('{$PUBLISH_DATE$}') >= 0:
                        cell.font = Font(name=u"ＭＳ 明朝", underline="single")
                    cell.value = replaced_string

    book.save(path)
    return path


def replace_excel_dict(sheet, detail):
    """エクセルの文字列を置換する。

    :param sheet: エクセルのワークシート
    :param detail 出力データ
    :return:
    """
    for key, value in detail.iteritems():
        sheet.Cells.Replace(What="{$%s$}" % (key,), Replacement=value, LookAt=constants.xlPart, MatchCase=False,
                            SearchFormat=False, ReplaceFormat=False, SearchOrder=constants.xlByRows)


def replace_excel_list(sheet, items, range_start='ITERATOR_START', range_end='ITERATOR_END'):
    """エクセルの繰り返す部分を置換する。

    :param sheet: エクセルのワークシート
    :param items 出力データ
    :param range_start:
    :param range_end:
    :return:
    """
    def get_positions(cell1, cell2):
        ret_value = []
        for sub_r in range(cell1.Row, cell2.Row + 1):
            for sub_c in range(cell1.Column, cell2.Column + 1):
                if sheet.Cells(sub_r, sub_c).Value:
                    ret_value.append((sub_r, sub_c, sheet.Cells(sub_r, sub_c).Value))
        return ret_value

    def get_row_span(sub_position):
        sub_rows = 0
        prev_r = 0
        for sub_r, sub_c, sub_text in sub_position:
            if not prev_r or prev_r != sub_r:
                prev_r = sub_r
                sub_rows += 1
        return sub_rows

    try:
        # 足りない行を追加する。
        start_cell = sheet.Range(range_start)
        end_cell = sheet.Range(range_end)
        positions = get_positions(start_cell, end_cell)
        if not items:
            for r, c, text in positions:
                sheet.Cells(r, c).Value = ""
            return

        row_span = get_row_span(positions)
        existed_rows = end_cell.row - start_cell.row + 1
        if len(items) > 1 and len(items) * row_span > existed_rows:
            # 行が足りない場合は追加する。
            cnt_extra_rows = len(items) * row_span - existed_rows
            sheet.Range("%s:%s" % (end_cell.Row + 1, end_cell.Row + cnt_extra_rows)).Insert(Shift=constants.xlDown)
        for i, item in enumerate(items):
            for r, c, text in positions:
                if text:
                    replacements = common.get_excel_replacements(text)
                    if replacements:
                        for replacement in replacements:
                            val = item.get(replacement, "{$" + replacement + "$}")
                            if not isinstance(val, str) and not isinstance(val, unicode):
                                val = str(val)
                            text = text.replace("{$" + replacement + "$}", val)
                    sheet.Cells(r + (i * row_span), c).Value = text
    except Exception as ex:
        print ex.message


def generate_attendance_format(user, template_path, project_members, lump_projects, year=None, month=None):
    """出勤情報をダウンロードする

    :param user: ログインユーザー
    :param template_path: テンプレートの格納場所
    :param project_members: 案件メンバー
    :param lump_projects: 一括案件
    :param year:
    :param month:
    :return: エクセルのバイナリ
    """
    book = px.load_workbook(template_path)
    sheet = book.get_sheet_by_name('Sheet1')

    start_row = constants.POS_ATTENDANCE_START_ROW
    count = project_members.count() + lump_projects.count()
    set_openpyxl_styles(sheet, 'B5:AE%s' % (start_row + count + 2,), 5)
    for i, project_member in enumerate(project_members):
        # NO
        sheet.cell(row=start_row, column=2).value = i + 1
        # 隠し項目（Project_member ID)
        sheet.cell(row=start_row, column=3).value = project_member.id
        # 社員番号
        sheet.cell(row=start_row, column=4).value = project_member.member.employee_id
        # 氏名
        sheet.cell(row=start_row, column=5).value = project_member.member.__unicode__()
        # 所在部署
        section = project_member.member.get_section()
        if section:
            sheet.cell(row=start_row, column=6).value = section.__unicode__()
        # 会社
        if project_member.member.company:
            sheet.cell(row=start_row, column=7).value = project_member.member.company.name
        else:
            # 他者技術者
            sheet.cell(row=start_row, column=7).value = project_member.member.subcontractor.name
        # 契約形態
        sheet.cell(row=start_row, column=9).value = project_member.member.get_member_type_display()
        # 案件名
        sheet.cell(row=start_row, column=10).value = project_member.project.name
        # 顧客名
        sheet.cell(row=start_row, column=11).value = project_member.project.client.name
        # 契約種類
        sheet.cell(row=start_row, column=12).value = u"一括" if project_member.project.is_lump else 'SES'

        # 出勤情報取得
        if len(project_member.current_attendance_set) == 1:
            attendance = project_member.current_attendance_set[0]
            date = datetime.date(int(attendance.year), int(attendance.month), 20)
            is_own = project_member.member.is_belong_to(user, date)
        elif project_member.project.is_reserve:
            attendance = models.MemberAttendance(project_member=project_member, year=year, month=month)
            date = datetime.date(int(attendance.year), int(attendance.month), 20)
            is_own = project_member.member.is_belong_to(user, date)
        else:
            attendance = None
            is_own = False
        if attendance and is_own:
            # 社会保険加入有無
            contract = attendance.get_contract()
            if contract and hasattr(contract, 'endowment_insurance') and contract.endowment_insurance == "1":
                sheet.cell(row=start_row, column=8).value = u"○"
            # 勤務時間
            sheet.cell(row=start_row, column=13).value = attendance.total_hours
            # 勤務に数
            sheet.cell(row=start_row, column=14).value = attendance.total_days
            # 深夜日数
            sheet.cell(row=start_row, column=15).value = attendance.night_days
            # 客先立替金
            sheet.cell(row=start_row, column=16).value = attendance.advances_paid_client
            # 立替金
            sheet.cell(row=start_row, column=17).value = attendance.advances_paid
            # 勤務交通費
            sheet.cell(row=start_row, column=18).value = attendance.traffic_cost

            # 請求情報取得
            if len(project_member.project_request_detail_set) == 1:
                request_detail = project_member.project_request_detail_set[0]
            else:
                request_detail = None
            if request_detail:
                # 売上（税込）
                sheet.cell(row=start_row, column=19).value = request_detail.get_all_price()
                # 売上（税抜）
                sheet.cell(row=start_row, column=20).value = request_detail.total_price
                # 売上（経費）
                sheet.cell(row=start_row, column=21).value = request_detail.expenses_price
            elif project_member.project.is_lump:
                # 売上（税込）
                sheet.cell(row=start_row, column=19).value = project_member.project.all_price_lump
                # 売上（税抜）
                sheet.cell(row=start_row, column=20).value = project_member.project.lump_amount
                # 売上（経費）
                sheet.cell(row=start_row, column=21).value = 0

            if request_detail or project_member.project.is_lump or project_member.project.is_reserve:
                # 月給
                sheet.cell(row=start_row, column=22).value = attendance.get_cost()
                # 手当
                sheet.cell(row=start_row, column=23).value = attendance.allowance
                # 深夜手当
                sheet.cell(row=start_row, column=24).value = attendance.get_night_allowance()
                # 残業／控除
                sheet.cell(row=start_row, column=25).value = attendance.get_overtime_cost()
                # 交通費(原価)
                sheet.cell(row=start_row, column=26).value = attendance.traffic_cost
                # 経費(原価)
                sheet.cell(row=start_row, column=27).value = attendance.expenses
        else:
            if len(project_member.prev_attendance_set) == 1:
                prev_attendance = project_member.prev_attendance_set[0]
                if prev_attendance.traffic_cost:
                    # 勤務交通費
                    sheet.cell(row=start_row, column=18).value = prev_attendance.traffic_cost
                    # # 交通費(原価)
                    # sheet.cell(row=start_row, column=26).value = prev_attendance.traffic_cost
                if prev_attendance.allowance:
                    # 手当
                    sheet.cell(row=start_row, column=23).value = prev_attendance.allowance

        start_row += 1
    # 一括案件
    for i, lump_project in enumerate(lump_projects):
        # NO
        sheet.cell(row=start_row, column=2).value = project_members.count() + 1 + i
        # 案件名
        sheet.cell(row=start_row, column=10).value = lump_project.name
        # 顧客名
        sheet.cell(row=start_row, column=11).value = lump_project.client.name
        # 契約種類
        sheet.cell(row=start_row, column=12).value = u"一括"
        project_request = lump_project.project_request_set[0] if lump_project.project_request_set else None
        if project_request:
            # 売上（税込）
            sheet.cell(row=start_row, column=19).value = project_request.amount
            # 売上（税抜）
            sheet.cell(row=start_row, column=20).value = project_request.turnover_amount
            # 売上（経費）
            sheet.cell(row=start_row, column=21).value = project_request.expenses_amount
        start_row += 1

    # 合計
    sheet.cell(row=start_row, column=18).value = "他社技術者の合計"
    sheet.cell(row=start_row, column=19).value = '=SUMIF(I5:I{0}, "=他社技術者", S5:S{0})'.format(count + 4)
    sheet.cell(row=start_row, column=20).value = '=SUMIF(I5:I{0}, "=他社技術者", T5:T{0})'.format(count + 4)
    sheet.cell(row=start_row, column=21).value = '=SUMIF(I5:I{0}, "=他社技術者", U5:U{0})'.format(count + 4)
    sheet.cell(row=start_row, column=22).value = '=SUMIF(I5:I{0}, "=他社技術者", V5:V{0})'.format(count + 4)
    sheet.cell(row=start_row, column=23).value = '=SUMIF(I5:I{0}, "=他社技術者", W5:W{0})'.format(count + 4)
    sheet.cell(row=start_row, column=24).value = '=SUMIF(I5:I{0}, "=他社技術者", X5:X{0})'.format(count + 4)
    sheet.cell(row=start_row, column=25).value = '=SUMIF(I5:I{0}, "=他社技術者", Y5:Y{0})'.format(count + 4)
    sheet.cell(row=start_row, column=26).value = '=SUMIF(I5:I{0}, "=他社技術者", Z5:Z{0})'.format(count + 4)
    sheet.cell(row=start_row, column=27).value = '=SUMIF(I5:I{0}, "=他社技術者", AA5:AA{0})'.format(count + 4)
    sheet.cell(row=start_row, column=28).value = '=SUMIF(I5:I{0}, "=他社技術者", AB5:AB{0})'.format(count + 4)
    sheet.cell(row=start_row, column=29).value = '=SUMIF(I5:I{0}, "=他社技術者", AC5:AC{0})'.format(count + 4)
    sheet.cell(row=start_row, column=30).value = '=SUMIF(I5:I{0}, "=他社技術者", AD5:AD{0})'.format(count + 4)
    sheet.cell(row=start_row, column=31).value = '=SUMIF(I5:I{0}, "=他社技術者", AE5:AE{0})'.format(count + 4)
    sheet.cell(row=start_row + 1, column=18).value = "自社の合計"
    sheet.cell(row=start_row + 1, column=19).value = '=SUMIF(I5:I{0}, "<>他社技術者", S5:S{0})'.format(count + 4)
    sheet.cell(row=start_row + 1, column=20).value = '=SUMIF(I5:I{0}, "<>他社技術者", T5:T{0})'.format(count + 4)
    sheet.cell(row=start_row + 1, column=21).value = '=SUMIF(I5:I{0}, "<>他社技術者", U5:U{0})'.format(count + 4)
    sheet.cell(row=start_row + 1, column=22).value = '=SUMIF(I5:I{0}, "<>他社技術者", V5:V{0})'.format(count + 4)
    sheet.cell(row=start_row + 1, column=23).value = '=SUMIF(I5:I{0}, "<>他社技術者", W5:W{0})'.format(count + 4)
    sheet.cell(row=start_row + 1, column=24).value = '=SUMIF(I5:I{0}, "<>他社技術者", X5:X{0})'.format(count + 4)
    sheet.cell(row=start_row + 1, column=25).value = '=SUMIF(I5:I{0}, "<>他社技術者", Y5:Y{0})'.format(count + 4)
    sheet.cell(row=start_row + 1, column=26).value = '=SUMIF(I5:I{0}, "<>他社技術者", Z5:Z{0})'.format(count + 4)
    sheet.cell(row=start_row + 1, column=27).value = '=SUMIF(I5:I{0}, "<>他社技術者", AA5:AA{0})'.format(count + 4)
    sheet.cell(row=start_row + 1, column=28).value = '=SUMIF(I5:I{0}, "<>他社技術者", AB5:AB{0})'.format(count + 4)
    sheet.cell(row=start_row + 1, column=29).value = '=SUMIF(I5:I{0}, "<>他社技術者", AC5:AC{0})'.format(count + 4)
    sheet.cell(row=start_row + 1, column=30).value = '=SUMIF(I5:I{0}, "<>他社技術者", AD5:AD{0})'.format(count + 4)
    sheet.cell(row=start_row + 1, column=31).value = '=SUMIF(I5:I{0}, "<>他社技術者", AE5:AE{0})'.format(count + 4)
    sheet.cell(row=start_row + 2, column=18).value = "全ての合計"
    sheet.cell(row=start_row + 2, column=19).value = "=SUM(S5:S%s)" % (count + 4)
    sheet.cell(row=start_row + 2, column=20).value = "=SUM(T5:T%s)" % (count + 4)
    sheet.cell(row=start_row + 2, column=21).value = "=SUM(U5:U%s)" % (count + 4)
    sheet.cell(row=start_row + 2, column=22).value = "=SUM(V5:V%s)" % (count + 4)
    sheet.cell(row=start_row + 2, column=23).value = "=SUM(W5:W%s)" % (count + 4)
    sheet.cell(row=start_row + 2, column=24).value = "=SUM(X5:X%s)" % (count + 4)
    sheet.cell(row=start_row + 2, column=25).value = "=SUM(Y5:Y%s)" % (count + 4)
    sheet.cell(row=start_row + 2, column=26).value = "=SUM(Z5:Z%s)" % (count + 4)
    sheet.cell(row=start_row + 2, column=27).value = "=SUM(AA5:AA%s)" % (count + 4)
    sheet.cell(row=start_row + 2, column=28).value = "=SUM(AB5:AB%s)" % (count + 4)
    sheet.cell(row=start_row + 2, column=29).value = "=SUM(AC5:AC%s)" % (count + 4)
    sheet.cell(row=start_row + 2, column=30).value = "=SUM(AD5:AD%s)" % (count + 4)
    sheet.cell(row=start_row + 2, column=31).value = "=SUM(AE5:AE%s)" % (count + 4)

    return save_virtual_workbook(book)


def generate_members_cost(user, members):
    book = px.Workbook()
    sheet = book.create_sheet(title=u"社員一覧", index=0)
    # タイトル
    for i, title in enumerate([u'給料王ＩＤ', u"名前", u"契約形態", u"メールアドレス",
                               u"事業部", u"部署", u"課・グループ", u"課長", u"協力会社", u"ランク",
                               u"月給", u"保険", u"残業", u"欠勤", u"案件"], start=1):
        sheet.cell(row=1, column=i).value = title
    # リスト
    start_row = 2
    today = datetime.date.today()
    for i, member in enumerate(members):
        contract = member.get_current_contract()
        section_period = member.current_section_period[0] if member.current_section_period else None
        is_own = member.is_belong_to(user, today)
        row = start_row + i
        sheet.cell(row=row, column=1).value = member.employee_id if contract and contract.member_type in (1, 2) else ''
        sheet.cell(row=row, column=2).value = unicode(member)
        sheet.cell(row=row, column=3).value = contract.get_member_type_display() if contract else ''
        sheet.cell(row=row, column=4).value = member.email
        sheet.cell(row=row, column=5).value = unicode(section_period.division) if section_period else ''
        sheet.cell(row=row, column=6).value = unicode(section_period.section) if section_period else ''
        sheet.cell(row=row, column=7).value = unicode(section_period.subsection) if section_period else ''
        sheet.cell(row=row, column=8).value = unicode(section_period.subsection.get_chief().first()) \
            if section_period and section_period.subsection else ''
        sheet.cell(row=row, column=9).value = unicode(member.subcontractor) if member.subcontractor else ''
        sheet.cell(row=row, column=10).value = member.get_ranking_display()
        if is_own:
            sheet.cell(row=row, column=11).value = contract.get_cost() if contract else ''
            sheet.cell(row=row, column=12).value = u"○" if contract and contract.endowment_insurance == "1" else ''
            sheet.cell(row=row, column=13).value = contract.allowance_overtime if contract else ''
            sheet.cell(row=row, column=14).value = contract.allowance_absenteeism if contract else ''
        else:
            sheet.cell(row=row, column=11).value = "****"
            sheet.cell(row=row, column=12).value = "****"
            sheet.cell(row=row, column=13).value = "****"
            sheet.cell(row=row, column=14).value = "****"
        project_member = member.get_current_project_member()
        sheet.cell(row=row, column=15).value = unicode(project_member.project) if project_member else ''

    return save_virtual_workbook(book)


def generate_eboa_members(members):
    book = px.Workbook()
    sheet = book.create_sheet(title=u"社員一覧", index=0)
    # タイトル
    for i, title in enumerate([u"ID", u'名前', u"フリカナ", u"雇用区分", u"所属部署",
                               u"性別", u"生年月日", u"郵便番号", u"住所", u"個人携帯番号", u"会社メールアドレス",
                               u"在留資格", u"在留カード番号", u"在留期間", u"入社年月日", u"雇用保険加入状況",
                               u"雇用保険番号", u"社会保険加入状況", u"社会保険番号", u"厚生年金加入状況",
                               u"厚生年金番号"], start=1):
        sheet.cell(row=1, column=i).value = title
    # １列目はＩＤなので、非表示にする。
    sheet.column_dimensions['A'].hidden = True
    # リスト
    start_row = 2
    for i, member in enumerate(members):
        contract = member.get_contract()
        organization = member.get_organization()
        row = start_row + i
        sheet.cell(row=row, column=1).value = member.pk
        sheet.cell(row=row, column=2).value = member.name
        sheet.cell(row=row, column=3).value = member.residence_name_kana or ''
        sheet.cell(row=row, column=4).value = contract.employer_type if contract else ''
        sheet.cell(row=row, column=5).value = organization.orgname if organization else ''
        sheet.cell(row=row, column=6).value = member.get_sex_display()
        sheet.cell(row=row, column=7).value = member.birthday
        sheet.cell(row=row, column=8).value = member.zipcode
        sheet.cell(row=row, column=9).value = member.address
        sheet.cell(row=row, column=10).value = member.private_tel_number
        sheet.cell(row=row, column=11).value = member.business_mail_addr
        sheet.cell(row=row, column=12).value = member.residence_type
        sheet.cell(row=row, column=13).value = member.id_number
        sheet.cell(row=row, column=14).value = member.id_card_expired_date
        sheet.cell(row=row, column=15).value = member.join_date
        sheet.cell(row=row, column=16).value = ''
        sheet.cell(row=row, column=17).value = ''
        sheet.cell(row=row, column=18).value = u"○" if contract and contract.endowment_insurance == "2" else ''
        sheet.cell(row=row, column=19).value = ''
        sheet.cell(row=row, column=20).value = ''
        sheet.cell(row=row, column=21).value = ''

    return save_virtual_workbook(book)


def set_openpyxl_styles(ws, cell_range, start_row):
    rows = list(ws.iter_rows(cell_range))
    style_list = []
    reg = re.compile(r'\b[A-Z]{1,2}([0-9])+\b')
    dict_formulae = {}

    # we convert iterator to list for simplicity, but it's not memory efficient solution
    rows = list(rows)
    for row_index, cells in enumerate(rows):
        for col_index, cell in enumerate(cells):
            if row_index == 0:
                style_list.append((cell.border.copy(),
                                   cell.font.copy(),
                                   cell.fill.copy(),
                                   cell.alignment.copy(),
                                   cell.number_format))
                # フォーミュラ
                if cell.value and cell.value[0] == '=':
                    lst = reg.findall(cell.value)
                    if lst and lst.count(lst[0]) == len(lst) and int(lst[0]) == start_row:
                        dict_formulae[col_index] = cell.value.replace(lst[0], '{0}')
            else:
                cell.border = style_list[col_index][0]
                cell.font = style_list[col_index][1]
                cell.fill = style_list[col_index][2]
                cell.alignment = style_list[col_index][3]
                cell.number_format = style_list[col_index][4]
                if cell.value and cell.value[0] == '=':
                    pass
                elif col_index in dict_formulae:
                    formulae = dict_formulae[col_index].format(start_row + row_index)
                    cell.value = formulae


def get_openpyxl_replaced_string(str_format, data):
    """openpyxlを利用して、置換文字列を置換する。

    :param str_format: {$XXXX$}を含む文字列
    :param data:
    :return:
    """
    match_list = re.findall(ur"\{\$[A-Z_0-9]+\$\}", str_format)
    if match_list:
        replaced_string = str_format
        for replace_item in match_list:
            replace_value = data.get(replace_item[2:-2], '')
            replaced_string = replaced_string.replace(replace_item, replace_value if replace_value else '')
        return replaced_string
    else:
        return None
